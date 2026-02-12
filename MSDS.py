import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage, ImageChops
import io
import re
import os
import fitz  # PyMuPDF
import numpy as np
import gc
import math
from datetime import datetime
import openpyxl.utils

# 1. 페이지 설정
st.set_page_config(page_title="MSDS 스마트 변환기", layout="wide")
st.title("MSDS 양식 변환기")
st.markdown("---")

# --------------------------------------------------------------------------
# [1. 유틸리티 함수]
# --------------------------------------------------------------------------
FONT_STYLE = Font(name='굴림', size=8)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

def safe_write_force(ws, row, col, value, center=False):
    cell = ws.cell(row=row, column=col)
    try: cell.value = value
    except AttributeError:
        try:
            for rng in list(ws.merged_cells.ranges):
                if cell.coordinate in rng:
                    ws.unmerge_cells(str(rng))
                    cell = ws.cell(row=row, column=col)
                    break
            cell.value = value
        except: pass
    if cell.font.name != '굴림': cell.font = FONT_STYLE
    if center: cell.alignment = ALIGN_CENTER
    else: cell.alignment = ALIGN_LEFT

def get_description_smart(code, code_map):
    clean_code = str(code).replace(" ", "").upper().strip()
    if clean_code in code_map: return code_map[clean_code]
    if "+" in clean_code:
        parts = clean_code.split("+")
        found_texts = []
        for p in parts:
            if p in code_map: found_texts.append(code_map[p])
        if found_texts: return " ".join(found_texts)
    return ""

def calculate_smart_height_basic(text): 
    if not text: return 19.2
    explicit_lines = str(text).count('\n') + 1
    final_lines = max(explicit_lines, 1)
    if final_lines == 1: return 19.2
    elif final_lines == 2: return 23.3
    else: return 33.0

def format_and_calc_height_sec47(text):
    if not text: return "", 19.2
    formatted_text = re.sub(r'(?<!\d)\.(?!\d)(?!\n)', '.\n', text)
    lines = [line.strip() for line in formatted_text.split('\n') if line.strip()]
    final_text = "\n".join(lines)
    char_limit_per_line = 45
    total_visual_lines = 0
    for line in lines:
        line_len = 0
        for ch in line:
            line_len += 2 if '가' <= ch <= '힣' else 1.1 
        visual_lines = math.ceil(line_len / (char_limit_per_line * 2)) 
        if visual_lines == 0: visual_lines = 1
        total_visual_lines += visual_lines
    if total_visual_lines == 0: total_visual_lines = 1
    height = (total_visual_lines * 10) + 10
    return final_text, height

def fill_fixed_range(ws, start_row, end_row, codes, code_map):
    unique_codes = []; seen = set()
    for c in codes:
        clean = c.replace(" ", "").upper().strip()
        if clean not in seen: unique_codes.append(clean); seen.add(clean)
    limit = end_row - start_row + 1
    for i in range(limit):
        current_row = start_row + i
        if i < len(unique_codes):
            code = unique_codes[i]
            desc = get_description_smart(code, code_map)
            ws.row_dimensions[current_row].hidden = False
            final_height = calculate_smart_height_basic(desc)
            ws.row_dimensions[current_row].height = final_height
            safe_write_force(ws, current_row, 2, code, center=False)
            safe_write_force(ws, current_row, 4, desc, center=False)
        else:
            ws.row_dimensions[current_row].hidden = True
            safe_write_force(ws, current_row, 2, "") 
            safe_write_force(ws, current_row, 4, "")

def fill_composition_data(ws, comp_data, cas_to_name_map, mode="CFF(K)"):
    start_row = 80; end_row = 123
    if "E" in mode: end_row = 122
    limit = end_row - start_row + 1

    for i in range(limit):
        current_row = start_row + i
        if i < len(comp_data):
            cas_no, concentration = comp_data[i]
            clean_cas = cas_no.replace(" ", "").strip()
            chem_name = cas_to_name_map.get(clean_cas, "")
            
            ws.row_dimensions[current_row].hidden = False
            ws.row_dimensions[current_row].height = 26.7
            
            # [K/E 공통] CAS NO 왼쪽 정렬 (center=False)
            safe_write_force(ws, current_row, 1, chem_name, center=False)
            safe_write_force(ws, current_row, 4, cas_no, center=False) 
            safe_write_force(ws, current_row, 6, concentration if concentration else "", center=True)
        else:
            ws.row_dimensions[current_row].hidden = True
            safe_write_force(ws, current_row, 1, "")
            safe_write_force(ws, current_row, 4, "")
            safe_write_force(ws, current_row, 6, "")

def fill_regulatory_section(ws, start_row, end_row, substances, data_map, col_key):
    limit = end_row - start_row + 1
    for i in range(limit):
        current_row = start_row + i
        if i < len(substances):
            substance_name = substances[i]
            safe_write_force(ws, current_row, 1, substance_name, center=False)
            cell_data = ""
            if substance_name in data_map:
                cell_data = str(data_map[substance_name].get(col_key, ""))
                if cell_data == "nan": cell_data = ""
            safe_write_force(ws, current_row, 2, cell_data, center=False)
            ws.row_dimensions[current_row].hidden = False
            _, h = format_and_calc_height_sec47(cell_data)
            if h < 24.0: h = 24.0 
            ws.row_dimensions[current_row].height = h
        else:
            safe_write_force(ws, current_row, 1, "")
            safe_write_force(ws, current_row, 2, "")
            ws.row_dimensions[current_row].hidden = True

# --------------------------------------------------------------------------
# [2. 이미지 함수]
# --------------------------------------------------------------------------
def auto_crop(pil_img):
    try:
        if pil_img.mode != 'RGB':
            bg = PILImage.new('RGB', pil_img.size, (255, 255, 255))
            if pil_img.mode == 'RGBA': bg.paste(pil_img, mask=pil_img.split()[3])
            else: bg.paste(pil_img)
            pil_img = bg
        bbox = ImageChops.invert(pil_img).getbbox()
        if bbox: return pil_img.crop(bbox)
        return pil_img
    except: return pil_img

def normalize_image_legacy(pil_img):
    try:
        if pil_img.mode in ('RGBA', 'LA') or (pil_img.mode == 'P' and 'transparency' in pil_img.info):
            background = PILImage.new('RGB', pil_img.size, (255, 255, 255))
            if pil_img.mode == 'P': pil_img = pil_img.convert('RGBA')
            background.paste(pil_img, mask=pil_img.split()[3])
            pil_img = background
        else: pil_img = pil_img.convert('RGB')
        return pil_img.resize((32, 32)).convert('L')
    except: return pil_img.resize((32, 32)).convert('L')

def normalize_image_smart(pil_img):
    try:
        cropped_img = auto_crop(pil_img)
        return cropped_img.resize((64, 64)).convert('L')
    except: return pil_img.resize((64, 64)).convert('L')

def get_reference_images():
    img_folder = "reference_imgs"
    if not os.path.exists(img_folder): return {}, False
    try:
        ref_images = {}
        file_list = sorted(os.listdir(img_folder)) 
        for fname in file_list:
            if fname.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.tif', '.tiff')):
                full_path = os.path.join(img_folder, fname)
                try:
                    pil_img = PILImage.open(full_path)
                    ref_images[fname] = pil_img
                except: continue
        return ref_images, True
    except: return {}, False

def is_blue_dominant(pil_img):
    try:
        img = pil_img.resize((50, 50)).convert('RGB')
        data = np.array(img)
        r = data[:,:,0].astype(int); g = data[:,:,1].astype(int); b = data[:,:,2].astype(int)
        blue_mask = (b > r + 30) & (b > g + 30)
        blue_ratio = np.sum(blue_mask) / (50 * 50)
        return blue_ratio > 0.05
    except: return False

def is_square_shaped(width, height):
    if height == 0: return False
    ratio = width / height
    return 0.8 < ratio < 1.2

def find_best_match_name(src_img, ref_images, mode="CFF(K)"):
    best_score = float('inf')
    best_name = None
    
    # [수정] 이미지 매칭 임계값 조정
    if mode == "HP(K)" or mode == "HP(E)":
        src_norm = normalize_image_smart(src_img)
        threshold = 60 # [수정] 80 -> 60으로 하향 조정 (이미지 매칭 완화)
    else: 
        src_norm = normalize_image_legacy(src_img)
        threshold = 60

    try:
        src_arr = np.array(src_norm, dtype='int16')
        for name, ref_img in ref_images.items():
            if "HP" in mode: ref_norm = normalize_image_smart(ref_img)
            else: ref_norm = normalize_image_legacy(ref_img)
                
            ref_arr = np.array(ref_norm, dtype='int16')
            diff = np.mean(np.abs(src_arr - ref_arr))
            if diff < best_score:
                best_score = diff
                best_name = name
        
        if best_score < threshold: return best_name, best_score
        else: return None, None
    except: return None, None

def extract_number(filename):
    nums = re.findall(r'\d+', filename)
    return int(nums[0]) if nums else 999

# --------------------------------------------------------------------------
# [3. 파서 함수]
# --------------------------------------------------------------------------
def get_clustered_lines(doc):
    all_lines = []
    noise_regexs = [
        r'^\s*\d+\s*/\s*\d+\s*$', r'물질안전보건자료', r'Material Safety Data Sheet', 
        r'PAGE', r'Ver\.\s*:?\s*\d+\.?\d*', r'발행일\s*:?.*', r'Date of issue',
        r'주식회사\s*고려.*', r'Cff', r'Corea\s*flavors.*', r'제\s*품\s*명\s*:?.*',
        r'according to the Global Harmonized System', r'Product Name', 
        r'Date\s*:\s*\d{2}\.[a-zA-Z]{3}\.\d{4}'
    ]
    global_y_offset = 0
    for page in doc:
        page_h = page.rect.height
        clip_rect = fitz.Rect(0, 60, page.rect.width, page_h - 50)
        words = page.get_text("words", clip=clip_rect)
        words.sort(key=lambda w: w[1]) 
        rows = []
        if words:
            current_row = [words[0]]
            row_base_y = words[0][1]
            for w in words[1:]:
                if abs(w[1] - row_base_y) < 8:
                    current_row.append(w)
                else:
                    current_row.sort(key=lambda x: x[0])
                    rows.append(current_row)
                    current_row = [w]
                    row_base_y = w[1]
            if current_row:
                current_row.sort(key=lambda x: x[0])
                rows.append(current_row)
        for row in rows:
            line_text = " ".join([w[4] for w in row])
            is_noise = False
            for pat in noise_regexs:
                if re.search(pat, line_text, re.IGNORECASE):
                    is_noise = True; break
            if not is_noise:
                avg_y = sum([w[1] for w in row]) / len(row)
                all_lines.append({
                    'text': line_text,
                    'global_y0': avg_y + global_y_offset,
                    'global_y1': (sum([w[3] for w in row]) / len(row)) + global_y_offset
                })
        global_y_offset += page_h
    return all_lines

def extract_section_smart(all_lines, start_kw, end_kw, mode="CFF(K)"):
    start_idx = -1; end_idx = -1
    clean_start_kw = start_kw.replace(" ", "")
    for i, line in enumerate(all_lines):
        if "E" in mode:
            if clean_start_kw.lower() in line['text'].replace(" ", "").lower():
                start_idx = i; break
        else:
            if clean_start_kw in line['text'].replace(" ", ""):
                start_idx = i; break
    if start_idx == -1: return ""
    
    if isinstance(end_kw, str): end_kw = [end_kw]
    clean_end_kws = [k.replace(" ", "") for k in end_kw]
    
    for i in range(start_idx + 1, len(all_lines)):
        line_clean = all_lines[i]['text'].replace(" ", "")
        for cek in clean_end_kws:
            if "E" in mode:
                if cek.lower() in line_clean.lower(): end_idx = i; break
            else:
                if cek in line_clean: end_idx = i; break
        if end_idx != -1: break
    if end_idx == -1: end_idx = len(all_lines)
    
    target_lines_raw = all_lines[start_idx : end_idx]
    if not target_lines_raw: return ""
    
    first_line = target_lines_raw[0].copy()
    txt = first_line['text']
    escaped_kw = re.escape(start_kw)
    pattern_str = escaped_kw.replace(r"\ ", r"\s*")
    match = re.search(pattern_str, txt, re.IGNORECASE)
    if match:
        content_part = txt[match.end():].strip()
        content_part = re.sub(r"^[:\.\-\s]+", "", content_part)
        first_line['text'] = content_part
    else:
        if start_kw in txt:
            parts = txt.split(start_kw, 1)
            first_line['text'] = parts[1].strip() if len(parts) > 1 else ""
        else: first_line['text'] = ""
    
    target_lines = []; 
    if first_line['text'].strip(): target_lines.append(first_line)
    target_lines.extend(target_lines_raw[1:])
    if not target_lines: return ""
    
    if mode == "CFF(E)":
        garbage_heads = [
            "Classification of the substance or mixture", "Classification of the substance or", "mixture",
            "Precautionary statements", "Hazard pictograms", "Signal word", 
            "Hazard statements", "Response", "Storage", "Disposal", "Other hazards",
            "General advice", "In case of eye contact", "In case of skin contact", "If inhaled", "If swallowed",
            "Special note for doctors", "Extinguishing media", "Special hazards arising from the", "Advice for firefighters",
            "Personal precautions, protective", "Environmental precautions", "Methods and materials for containment",
            "Precautions for safe handling", "Conditions for safe storage, including",
            "Internal regulations", "ACGIH regulations", "Biological exposure standards"
        ]
        sensitive_garbage_regex = []
    elif mode == "HP(K)":
        garbage_heads = ["에 접촉했을 때", "에 들어갔을 때", "들어갔을 때", "접촉했을 때", "했을 때", "흡입했을 때", "먹었을 때", "주의사항", "내용물", "취급요령", "저장방법", "보호구", "조치사항", "제거 방법", "소화제", "유해성", "로부터 생기는", "착용할 보호구", "예방조치", "방법", "경고표지 항목", "그림문자", "화학물질", "의사의 주의사항", "기타 의사의 주의사항", "필요한 정보", "관한 정보", "보호하기 위해 필요한 조치사항", "또는 제거 방법", "시 착용할 보호구 및 예방조치", "시 착용할 보호구", "부터 생기는 특정 유해성", "사의 주의사항", "(부적절한) 소화제", "및", "요령", "때", "항의", "색상", "인화점", "비중", "굴절률", "에 의한 규제", "의한 규제", "- 색", "(및 부적절한) 소화제", "특정 유해성", "보호하기 위해 필요한 조치 사항 및 보호구", "저장 방법"]
        sensitive_garbage_regex = [r"^시\s+", r"^또는\s+", r"^의\s+"]
    else: 
        garbage_heads = ["에 접촉했을 때", "에 들어갔을 때", "들어갔을 때", "접촉했을 때", "했을 때", "흡입했을 때", "먹었을 때", "주의사항", "내용물", "취급요령", "저장방법", "보호구", "조치사항", "제거 방법", "소화제", "유해성", "로부터 생기는", "착용할 보호구", "예방조치", "방법", "경고표지 항목", "그림문자", "화학물질", "의사의 주의사항", "기타 의사의 주의사항", "필요한 정보", "관한 정보", "보호하기 위해 필요한 조치사항", "또는 제거 방법", "시 착용할 보호구 및 예방조치", "시 착용할 보호구", "부터 생기는 특정 유해성", "사의 주의사항", "(부적절한) 소화제", "및", "요령", "때", "항의", "색상", "인화점", "비중", "굴절률", "에 의한 규제", "의한 규제"]
        sensitive_garbage_regex = [r"^시\s+", r"^또는\s+", r"^의\s+"]

    cleaned_lines = []
    for line in target_lines:
        txt = line['text'].strip()
        if "HP" in mode: txt = txt.lstrip("-").strip()
        
        for _ in range(3):
            changed = False
            for gb in garbage_heads:
                if txt.lower().replace(" ","").startswith(gb.lower().replace(" ","")):
                      p = re.compile(r"^" + re.escape(gb).replace(r"\ ", r"\s*") + r"[\s\.:]*", re.IGNORECASE)
                      m = p.match(txt)
                      if m: txt = txt[m.end():].strip(); changed = True
                      elif txt.lower().startswith(gb.lower()): txt = txt[len(gb):].strip(); changed = True
            
            for pat in sensitive_garbage_regex:
                m = re.search(pat, txt)
                if m: txt = txt[m.end():].strip(); changed = True
            
            txt = re.sub(r"^[:\.\)\s]+", "", txt)
            if not changed: break
        
        if txt:
            if "HP" in mode: txt = txt.lstrip("-").strip()
            line['text'] = txt
            cleaned_lines.append(line)
    
    if not cleaned_lines: return ""

    final_text = ""
    if len(cleaned_lines) > 0:
        final_text = cleaned_lines[0]['text']
        for i in range(1, len(cleaned_lines)):
            prev = cleaned_lines[i-1]; curr = cleaned_lines[i]
            if mode == "CFF(E)":
                final_text += "\n" + curr['text']
            else: 
                prev_txt = prev['text'].strip(); curr_txt = curr['text'].strip()
                ends_with_sentence = re.search(r"(\.|시오|음|함|것|임|있음|주의|금지|참조|따르시오|마시오)$", prev_txt)
                starts_with_bullet = re.match(r"^(\-|•|\*|\d+\.|[가-하]\.|\(\d+\))", curr_txt)
                if ends_with_sentence or starts_with_bullet: final_text += "\n" + curr_txt
                else:
                    last_char = prev_txt[-1] if prev_txt else ""
                    first_char = curr_txt[0] if curr_txt else ""
                    is_last_hangul = 0xAC00 <= ord(last_char) <= 0xD7A3
                    is_first_hangul = 0xAC00 <= ord(first_char) <= 0xD7A3
                    gap = curr['global_y0'] - prev['global_y1']
                    if gap < 3.0: 
                        if is_last_hangul and is_first_hangul:
                            need_space = False
                            if last_char in ['을', '를', '이', '가', '은', '는', '의', '와', '과', '에', '로', '서']: need_space = True
                            elif last_char in ['고', '며', '여', '해', '나', '면', '니', '등', '및', '또는', '경우', ',', ')', '속']: need_space = True
                            elif any(curr_txt.startswith(x) for x in ['및', '또는', '(', '참고']): need_space = True
                            if need_space: final_text += " " + curr_txt
                            else: final_text += curr_txt
                        else: final_text += " " + curr_txt
                    else: final_text += "\n" + curr_txt
    return final_text

def parse_sec8_hp_content(text):
    if not text: return "자료없음"
    chunks = text.split("-")
    valid_lines = []
    for chunk in chunks:
        clean_chunk = chunk.strip()
        if not clean_chunk: continue
        if ":" in clean_chunk:
            parts = clean_chunk.split(":", 1)
            name_part = parts[0].strip()
            value_part = parts[1].strip()
            if "해당없음" in value_part: continue 
            name_part = name_part.replace("[", "").replace("]", "").strip()
            value_part = value_part.replace("[", "").replace("]", "").strip()
            final_line = f"{name_part} : {value_part}"
            valid_lines.append(final_line)
        else:
            if "해당없음" not in clean_chunk:
                clean_chunk = clean_chunk.replace("[", "").replace("]", "").strip()
                valid_lines.append(clean_chunk)
    if not valid_lines: return "자료없음"
    return "\n".join(valid_lines)

# --------------------------------------------------------------------------
# [메인 파서]
# --------------------------------------------------------------------------
def parse_pdf_final(doc, mode="CFF(K)"):
    all_lines = get_clustered_lines(doc)
    
    result = {
        "hazard_cls": [], "signal_word": "", "h_codes": [], 
        "p_prev": [], "p_resp": [], "p_stor": [], "p_disp": [],
        "composition_data": [], "sec4_to_7": {}, "sec8": {}, "sec9": {}, "sec14": {}, "sec15": {}
    }
    
    if mode == "CFF(E)":
        hazard_cls_text = extract_section_smart(all_lines, "2. Hazards identification", "2.2 Labelling", mode)
        hazard_cls_lines = []
        for line in hazard_cls_text.split('\n'):
            line = line.strip()
            if not line: continue
            if "2.1 Classification" in line: continue
            if line.lower() == "mixture": continue
            hazard_cls_lines.append(line)
        result["hazard_cls"] = hazard_cls_lines

        full_text = "\n".join([l['text'] for l in all_lines])
        m_sig = re.search(r"Signal word\s*[:\-\s]*([A-Za-z]+)", full_text, re.IGNORECASE)
        if m_sig: result["signal_word"] = m_sig.group(1).capitalize()
        
        h_text = extract_section_smart(all_lines, "Hazard statements", "Precautionary statements", mode)
        result["h_codes"] = list(set(re.findall(r'\b(H\d{3}[a-zA-Z]*)\b', h_text)))

        p_prev_text = extract_section_smart(all_lines, "Precautionary statements", "Response", mode)
        result["p_prev"] = list(set(re.findall(r'\b(P\d{3}[a-zA-Z]*)\b', p_prev_text)))
        
        p_resp_text = extract_section_smart(all_lines, "Response", "Storage", mode)
        result["p_resp"] = list(set(re.findall(r'\b(P\d{3}[a-zA-Z]*)\b', p_resp_text)))
        
        p_stor_text = extract_section_smart(all_lines, "Storage", "Disposal", mode)
        result["p_stor"] = list(set(re.findall(r'\b(P\d{3}[a-zA-Z]*)\b', p_stor_text)))
        
        p_disp_text = extract_section_smart(all_lines, "Disposal", "2.3 Other hazards", mode)
        result["p_disp"] = list(set(re.findall(r'\b(P\d{3}[a-zA-Z]*)\b', p_disp_text)))

        comp_text = extract_section_smart(all_lines, "3. Composition", "4. FIRST-AID", mode)
        regex_cas = re.compile(r'\b\d{2,7}-\d{2}-\d\b')
        regex_conc = re.compile(r'(\d+(?:\.\d+)?)\s*~\s*(\d+(?:\.\d+)?)')
        comp_lines = comp_text.split('\n')
        for line in comp_lines:
            cas_m = regex_cas.search(line)
            conc_m = regex_conc.search(line)
            if cas_m:
                c_val = cas_m.group(0)
                cn_val = ""
                if conc_m: cn_val = f"{conc_m.group(1)} ~ {conc_m.group(2)}"
                result["composition_data"].append((c_val, cn_val))

        data = {}
        data["B125"] = extract_section_smart(all_lines, "4.1 General advice", "4.2 In case of eye contact", mode)
        data["B126"] = extract_section_smart(all_lines, "4.2 In case of eye contact", "4.3 In case of skin contact", mode)
        data["B127"] = extract_section_smart(all_lines, "4.3 In case of skin contact", "4.4 If inhaled", mode)
        data["B128"] = extract_section_smart(all_lines, "4.4 If inhaled", "4.5 If swallowed", mode)
        data["B129"] = extract_section_smart(all_lines, "4.5 If swallowed", "4.6 Special note for doctors", mode)
        if data["B129"]:
            data["B129"] = data["B129"].replace("Medical personnel, and to ensure that take protection measures is recognized for its substance", "")

        data["B132"] = extract_section_smart(all_lines, "5.1 Extinguishing media", "5.2 Special hazards", mode)
        data["B134"] = extract_section_smart(all_lines, "5.2 Special hazards", "5.3 Advice for firefighters", mode)
        if data["B134"]: data["B134"] = data["B134"].replace("substance or mixture", "")

        data["B136"] = extract_section_smart(all_lines, "5.3 Advice for firefighters", "6. Accidental", mode)
        data["B140"] = extract_section_smart(all_lines, "6.1 Personal precautions", "6.2 Environmental", mode)
        if data["B140"]: data["B140"] = data["B140"].replace("equipment and emergency procedures", "")

        data["B142"] = extract_section_smart(all_lines, "6.2 Environmental", "6.3 Methods", mode)
        data["B144"] = extract_section_smart(all_lines, "6.3 Methods", "7. Handling", mode)
        if data["B144"]: data["B144"] = data["B144"].replace("and cleaning up", "")

        data["B148"] = extract_section_smart(all_lines, "7.1 Precautions", "7.2 Conditions", mode)
        data["B150"] = extract_section_smart(all_lines, "7.2 Conditions", "8. Exposure", mode)
        if data["B150"]: data["B150"] = data["B150"].replace("any incompatibilities", "")

        result["sec4_to_7"] = data

        s8 = {}
        s8["B154"] = extract_section_smart(all_lines, "Internal regulations", "ACGIH regulations", mode)
        s8["B156"] = extract_section_smart(all_lines, "ACGIH regulations", "Biological exposure", mode)
        result["sec8"] = s8

        s9 = {}
        s9["B170"] = extract_section_smart(all_lines, "Color", "Odor", mode)
        s9["B176"] = extract_section_smart(all_lines, "Flash point", "Evaporation rate", mode)
        s9["B183"] = extract_section_smart(all_lines, "Specific gravity", "Partition coefficient", mode)
        s9["B189"] = extract_section_smart(all_lines, "Refractive index", "10. Stability", mode)
        result["sec9"] = s9

        s14 = {}
        un_text = extract_section_smart(all_lines, "14.1 UN number", "14.2 Proper", mode)
        s14["UN"] = re.sub(r'\D', '', un_text)
        name_text = extract_section_smart(all_lines, "14.2 Proper", "14.3 Transport", mode)
        s14["NAME"] = name_text
        result["sec14"] = s14

        return result

    if mode == "CFF(K)":
        for i in range(len(all_lines)):
            if "적정선적명" in all_lines[i]['text']:
                target_line = all_lines[i]
                if i > 0:
                    prev_line = all_lines[i-1]
                    if abs(prev_line['global_y0'] - target_line['global_y0']) < 20:
                        if "적정선적명" not in prev_line['text'] and "유엔번호" not in prev_line['text']:
                            all_lines[i]['text'] = target_line['text'] + " " + prev_line['text']
                            all_lines[i-1]['text'] = ""
    
    limit_y = 999999
    for line in all_lines:
        if "3. 구성성분" in line['text'] or "3. 성분" in line['text']:
            limit_y = line['global_y0']; break
    
    full_text_hp = "\n".join([l['text'] for l in all_lines if l['global_y0'] < limit_y])
    
    signal_found = False
    if mode == "HP(K)":
        try:
            start_sig = full_text_hp.find("신호어")
            end_sig = full_text_hp.find("유해", start_sig)
            if start_sig != -1 and end_sig != -1:
                target_area = full_text_hp[start_sig:end_sig]
                m = re.search(r"[-•]\s*(위험|경고)", target_area)
                if m: result["signal_word"] = m.group(1); signal_found = True
        except: pass
    if not signal_found:
        for line in full_text_hp.split('\n'):
            if "신호어" in line:
                val = line.replace("신호어", "").replace(":", "").strip()
                if val in ["위험", "경고"]: result["signal_word"] = val
            elif line.strip() in ["위험", "경고"] and not result["signal_word"]:
                result["signal_word"] = line.strip()
    
    if mode == "HP(K)":
        lines_hp = full_text_hp.split('\n')
        state = 0
        for l in lines_hp:
            if "가. 유해성" in l: state=1; continue
            if "나. 예방조치" in l: state=0; continue
            if state==1 and l.strip():
                if "공급자" not in l and "회사명" not in l:
                    clean_l = l.replace("-", "").strip()
                    if clean_l: result["hazard_cls"].append(clean_l)
    else: 
        lines_hp = full_text_hp.split('\n')
        state = 0
        for l in lines_hp:
            l_ns = l.replace(" ", "")
            if "2.유해성" in l_ns and "위험성" in l_ns: state = 1; continue 
            if "나.예방조치" in l_ns: state = 0; continue
            if state == 1 and l.strip():
                if "가.유해성" in l_ns and "분류" in l_ns:
                    check_header = re.sub(r'[가-하][\.\s]*유해성[\s\.]*위험성[\s\.]*분류', '', l).strip()
                    if not check_header: continue 
                    l = check_header
                if "공급자" not in l and "회사명" not in l:
                    result["hazard_cls"].append(l.strip())

    regex_code = re.compile(r"([HP]\s?\d{3}(?:\s*\+\s*[HP]\s?\d{3})*)")
    all_matches = regex_code.findall(full_text_hp)
    seen = set()
    if "P321" in full_text_hp and "P321" not in all_matches: all_matches.append("P321")
    for code_raw in all_matches:
        code = code_raw.replace(" ", "").upper()
        if code in seen: continue
        seen.add(code)
        if code.startswith("H"): result["h_codes"].append(code)
        elif code.startswith("P"):
            p = code.split("+")[0]
            if p.startswith("P2"): result["p_prev"].append(code)
            elif p.startswith("P3"): result["p_resp"].append(code)
            elif p.startswith("P4"): result["p_stor"].append(code)
            elif p.startswith("P5"): result["p_disp"].append(code)

    regex_conc = re.compile(r'\b(\d+(?:\.\d+)?)\s*(?:~|-)\s*(\d+(?:\.\d+)?)\b')
    regex_cas_strict = re.compile(r'\b(\d{2,7}\s*-\s*\d{2}\s*-\s*\d)\b')
    regex_cas_ec_kill = re.compile(r'\b\d{2,7}\s*-\s*\d{2,3}\s*-\s*\d\b')
    regex_tilde_range = re.compile(r'(\d+(?:\.\d+)?)\s*~\s*(\d+(?:\.\d+)?)') 
    
    in_comp = False
    for line in all_lines:
        txt = line['text']
        if "3." in txt and ("성분" in txt or "Composition" in txt): in_comp=True; continue
        if "4." in txt and ("응급" in txt or "First" in txt): in_comp=False; break
        if in_comp:
            if re.search(r'^\d+\.\d+', txt): continue 
            
            c_val = ""
            cn_val = ""
            
            if mode == "HP(K)":
                cas_found = regex_cas_strict.findall(txt)
                if cas_found:
                    c_val = cas_found[0].replace(" ", "")
                    txt_no_cas = txt.replace(cas_found[0], " " * len(cas_found[0]))
                    m_range = re.search(r'\b(\d+(?:\.\d+)?)\s*(?:-|~)\s*(\d+(?:\.\d+)?)\b', txt_no_cas)
                    if m_range:
                        s, e = m_range.group(1), m_range.group(2)
                        if s == "1": s = "0"
                        cn_val = f"{s} ~ {e}"
                    else:
                        m_single = re.search(r'\b(\d+(?:\.\d+)?)\b', txt_no_cas)
                        if m_single:
                            try:
                                if float(m_single.group(1)) <= 100: cn_val = m_single.group(1)
                            except: pass
            else:
                cas_found = regex_cas_ec_kill.findall(txt)
                if cas_found:
                    potential_cas = cas_found[0].replace(" ", "")
                    if re.match(r'\d{2,7}-\d{2}-\d', potential_cas): c_val = potential_cas
                txt_clean = regex_cas_ec_kill.sub(" ", txt)
                m_tilde = regex_tilde_range.search(txt_clean)
                if m_tilde:
                    s, e = m_tilde.group(1), m_tilde.group(2)
                    if s == "1": s = "0"
                    cn_val = f"{s} ~ {e}"
            
            if c_val or cn_val:
                if "." in cn_val: continue
                result["composition_data"].append((c_val, cn_val))

    data = {}
    if mode == "HP(K)":
        data["B125"] = extract_section_smart(all_lines, "가. 눈에", "나. 피부", mode)
        data["B126"] = extract_section_smart(all_lines, "나. 피부", "다. 흡입", mode)
        data["B127"] = extract_section_smart(all_lines, "다. 흡입", "라. 먹었을", mode)
        data["B128"] = extract_section_smart(all_lines, "라. 먹었을", "마. 기타", mode)
        data["B129"] = extract_section_smart(all_lines, "마. 기타", ["5.", "폭발"], mode)
        data["B132"] = extract_section_smart(all_lines, "가. 적절한", "나. 화학물질", mode)
        data["B133"] = extract_section_smart(all_lines, "나. 화학물질", "다. 화재진압", mode)
        data["B134"] = extract_section_smart(all_lines, "다. 화재진압", ["6.", "누출"], mode)
    else: 
        data["B125"] = extract_section_smart(all_lines, "나. 눈", "다. 피부", mode)
        data["B126"] = extract_section_smart(all_lines, "다. 피부", "라. 흡입", mode)
        data["B127"] = extract_section_smart(all_lines, "라. 흡입", "마. 먹었을", mode)
        data["B128"] = extract_section_smart(all_lines, "마. 먹었을", "바. 기타", mode)
        data["B129"] = extract_section_smart(all_lines, "바. 기타", ["5.", "폭발"], mode)
        data["B132"] = extract_section_smart(all_lines, "가. 적절한", "나. 화학물질", mode)
        data["B133"] = extract_section_smart(all_lines, "나. 화학물질", "다. 화재진압", mode)
        data["B134"] = extract_section_smart(all_lines, "다. 화재진압", ["6.", "누출"], mode)
    
    data["B138"] = extract_section_smart(all_lines, "가. 인체를", "나. 환경을", mode)
    data["B139"] = extract_section_smart(all_lines, "나. 환경을", "다. 정화", mode)
    data["B140"] = extract_section_smart(all_lines, "다. 정화", ["7.", "취급"], mode)
    data["B143"] = extract_section_smart(all_lines, "가. 안전취급", "나. 안전한", mode)
    data["B144"] = extract_section_smart(all_lines, "나. 안전한", ["8.", "노출"], mode)
    result["sec4_to_7"] = data

    sec8_lines = []
    start_8 = -1; end_8 = -1
    for i, line in enumerate(all_lines):
        if "8. 노출방지" in line['text']: start_8 = i
        if "9. 물리화학" in line['text']: end_8 = i; break
    if start_8 != -1:
        if end_8 == -1: end_8 = len(all_lines)
        sec8_lines = all_lines[start_8:end_8]
    
    if mode == "HP(K)":
        b148_raw = extract_section_smart(sec8_lines, "국내노출기준", "ACGIH노출기준", mode)
        b150_raw = extract_section_smart(sec8_lines, "ACGIH노출기준", "생물학적", mode)
        b148_raw = parse_sec8_hp_content(b148_raw)
        b150_raw = parse_sec8_hp_content(b150_raw)
    else:
        b148_raw = extract_section_smart(sec8_lines, "국내규정", "ACGIH", mode)
        b150_raw = extract_section_smart(sec8_lines, "ACGIH", "생물학적", mode)
    result["sec8"] = {"B148": b148_raw, "B150": b150_raw}

    sec9_lines = []
    start_9 = -1; end_9 = -1
    for i, line in enumerate(all_lines):
        if "9. 물리화학" in line['text']: start_9 = i
        if "10. 안정성" in line['text']: end_9 = i; break
    if start_9 != -1:
        if end_9 == -1: end_9 = len(all_lines)
        sec9_lines = all_lines[start_9:end_9]
        
    if mode == "HP(K)":
        result["sec9"] = {
            "B163": extract_section_smart(sec9_lines, "- 색", "나. 냄새", mode),
            "B169": extract_section_smart(sec9_lines, "인화점", "아. 증발속도", mode),
            "B176": extract_section_smart(sec9_lines, "비중", "거. n-옥탄올", mode),
            "B182": extract_section_smart(sec9_lines, "굴절률", ["10. 안정성", "10. 화학적"], mode)
        }
    else:
        result["sec9"] = {
            "B163": extract_section_smart(sec9_lines, "색상", "나. 냄새", mode),
            "B169": extract_section_smart(sec9_lines, "인화점", "아. 증발속도", mode),
            "B176": extract_section_smart(sec9_lines, "비중", "거. n-옥탄올", mode),
            "B182": extract_section_smart(sec9_lines, "굴절률", ["10. 안정성", "10. 화학적"], mode)
        }

    sec14_lines = []
    start_14 = -1; end_14 = -1
    for i, line in enumerate(all_lines):
        if "14. 운송에" in line['text']: start_14 = i
        if "15. 법적규제" in line['text']: end_14 = i; break
    if start_14 != -1:
        if end_14 == -1: end_14 = len(all_lines)
        sec14_lines = all_lines[start_14:end_14]
    
    if mode == "HP(K)":
        un_no = extract_section_smart(sec14_lines, "유엔번호", "나. 유엔", mode)
        ship_name = extract_section_smart(sec14_lines, "유엔 적정 선적명", ["다. 운송에서의", "다.운송에서의"], mode)
    else:
        un_no = extract_section_smart(sec14_lines, "유엔번호", "나. 적정선적명", mode)
        ship_name = extract_section_smart(sec14_lines, "적정선적명", ["다. 운송에서의", "다.운송에서의"], mode)
    result["sec14"] = {"UN": un_no, "NAME": ship_name}

    sec15_lines = []
    start_15 = -1; end_15 = -1
    for i, line in enumerate(all_lines):
        if "15. 법적규제" in line['text']: start_15 = i
        if "16. 그 밖의" in line['text']: end_15 = i; break
    if start_15 != -1:
        if end_15 == -1: end_15 = len(all_lines)
        sec15_lines = all_lines[start_15:end_15]
    
    if mode == "HP(K)":
        danger_act = ""
    else:
        danger_act = extract_section_smart(sec15_lines, "위험물안전관리법", "마. 폐기물", mode)
    result["sec15"] = {"DANGER": danger_act}

    return result

# --------------------------------------------------------------------------
# [4. 메인 실행 구역]
# --------------------------------------------------------------------------
with st.expander("📂 필수 파일 업로드", expanded=True):
    col1, col2 = st.columns(2)
    with col1:
        master_data_file = st.file_uploader("1. 중앙 데이터 (ingredients...xlsx)", type="xlsx")
        loaded_refs, folder_exists = get_reference_images()
        if folder_exists and loaded_refs:
            st.success(f"✅ 기준 그림 {len(loaded_refs)}개 로드됨")
        elif not folder_exists:
            st.warning("⚠️ 'reference_imgs' 폴더 필요")

    with col2:
        template_file = st.file_uploader("2. 양식 파일 (GHS MSDS 양식)", type="xlsx")

product_name_input = st.text_input("제품명 입력")
option = st.selectbox("적용할 양식", ("CFF(K)", "CFF(E)", "HP(K)", "HP(E)"))
st.write("") 

col_left, col_center, col_right = st.columns([4, 2, 4])

if 'converted_files' not in st.session_state:
    st.session_state['converted_files'] = []
    st.session_state['download_data'] = {}

with col_left:
    st.subheader("3. 원본 파일 업로드")
    uploaded_files = st.file_uploader("원본 데이터(PDF)", type=["pdf"], accept_multiple_files=True)

with col_center:
    st.write("") ; st.write("") ; st.write("")
    
    if st.button("▶ 변환 시작", use_container_width=True):
        if uploaded_files and master_data_file and template_file:
            with st.spinner(f"{option} 모드로 변환 중..."):
                
                new_files = []
                new_download_data = {}
                
                code_map = {} 
                cas_name_map = {} 
                kor_data_map = {} 
                eng_data_map = {} 
                
                try:
                    xls = pd.ExcelFile(master_data_file)
                    
                    target_sheet = None
                    for sheet in xls.sheet_names:
                        if "위험" in sheet and "안전" in sheet: target_sheet = sheet; break
                    
                    # [수정 시작] 위험 안전문구 시트 처리 로직 변경
                    if target_sheet:
                        # 헤더 없이 읽어서 인덱스로 접근
                        df_code = pd.read_excel(master_data_file, sheet_name=target_sheet)
                        
                        # "K" 모드이면 B열(1), 아니면 C열(2) 사용
                        if "K" in option:
                            target_col_idx = 1
                        else:
                            target_col_idx = 2
                        
                        for _, row in df_code.iterrows():
                            # A열(0)은 코드
                            if pd.notna(row.iloc[0]):
                                code_key = str(row.iloc[0]).replace(" ","").upper().strip()
                                # 지정된 컬럼 값 가져오기
                                if len(row) > target_col_idx:
                                    val = row.iloc[target_col_idx]
                                    code_val = str(val).strip() if pd.notna(val) else ""
                                    code_map[code_key] = code_val
                    # [수정 끝]
                    
                    if "K" in option:
                        sheet_kor = None
                        for sheet in xls.sheet_names:
                            if "국문" in sheet: sheet_kor = sheet; break
                        if sheet_kor:
                            df_kor = pd.read_excel(master_data_file, sheet_name=sheet_kor)
                            for _, row in df_kor.iterrows():
                                val_cas = row.iloc[0]
                                val_name = row.iloc[1]
                                if pd.notna(val_cas):
                                    c = str(val_cas).replace(" ", "").strip()
                                    n = str(val_name).strip() if pd.notna(val_name) else ""
                                    cas_name_map[c] = n
                                    if n:
                                        kor_data_map[n] = {
                                            'F': row.iloc[5], 'G': row.iloc[6], 'H': row.iloc[7],
                                            'P': row.iloc[15], 'T': row.iloc[19], 'U': row.iloc[20], 'V': row.iloc[21]
                                        }
                    else: # E 모드 (CFF E 등)
                        sheet_eng = None
                        for sheet in xls.sheet_names:
                            if "영문" in sheet: sheet_eng = sheet; break
                        if sheet_eng:
                            df_eng = pd.read_excel(master_data_file, sheet_name=sheet_eng)
                            for _, row in df_eng.iterrows():
                                val_cas = row.iloc[0]
                                val_name = row.iloc[1]
                                if pd.notna(val_cas):
                                    c = str(val_cas).replace(" ", "").strip()
                                    n = str(val_name).strip() if pd.notna(val_name) else ""
                                    cas_name_map[c] = n
                                    if n:
                                        eng_data_map[n] = {
                                            'F': row.iloc[5], 'G': row.iloc[6], 'H': row.iloc[7],
                                            'P': row.iloc[15], 'Q': row.iloc[16], 
                                            'T': row.iloc[19], 'U': row.iloc[20], 'V': row.iloc[21]
                                        }

                except Exception as e:
                    st.error(f"데이터 로드 오류: {e}")

                for uploaded_file in uploaded_files:
                    try:
                        doc = fitz.open(stream=uploaded_file.read(), filetype="pdf")
                        parsed_data = parse_pdf_final(doc, mode=option)
                        
                        template_file.seek(0)
                        dest_wb = load_workbook(io.BytesIO(template_file.read()))
                        dest_ws = dest_wb.active

                        dest_wb.external_links = []
                        dest_ws._images = []

                        for row in dest_ws.iter_rows():
                            for cell in row:
                                if isinstance(cell, MergedCell): continue
                                if cell.column == 2 and cell.data_type == 'f': cell.value = ""

                        if option == "CFF(E)":
                            safe_write_force(dest_ws, 6, 2, product_name_input, center=True)
                            safe_write_force(dest_ws, 9, 2, product_name_input, center=True)
                            
                            if parsed_data["hazard_cls"]:
                                clean_cls = "\n".join(parsed_data["hazard_cls"])
                                safe_write_force(dest_ws, 19, 2, clean_cls, center=False)
                            
                            if parsed_data["signal_word"]:
                                safe_write_force(dest_ws, 23, 2, parsed_data["signal_word"], center=False)

                            fill_fixed_range(dest_ws, 24, 36, parsed_data["h_codes"], code_map)
                            fill_fixed_range(dest_ws, 38, 49, parsed_data["p_prev"], code_map)
                            fill_fixed_range(dest_ws, 50, 63, parsed_data["p_resp"], code_map)
                            fill_fixed_range(dest_ws, 64, 69, parsed_data["p_stor"], code_map)
                            fill_fixed_range(dest_ws, 70, 72, parsed_data["p_disp"], code_map)
                            
                            fill_composition_data(dest_ws, parsed_data["composition_data"], cas_name_map, mode=option)
                            
                            active_substances = []
                            for c_data in parsed_data["composition_data"]:
                                cas = c_data[0].replace(" ", "").strip()
                                if cas in cas_name_map:
                                    name = cas_name_map[cas]
                                    if name: active_substances.append(name)
                                    
                            sd = parsed_data["sec4_to_7"]
                            cell_map_e = {
                                "B125": sd.get("B125",""), "B126": sd.get("B126",""), 
                                "B127": sd.get("B127",""), "B128": sd.get("B128",""),
                                "B129": sd.get("B129",""), "B132": sd.get("B132",""),
                                "B134": sd.get("B134",""), "B136": sd.get("B136",""),
                                "B140": sd.get("B140",""), "B142": sd.get("B142",""),
                                "B144": sd.get("B144",""), "B148": sd.get("B148",""),
                                "B150": sd.get("B150",""),
                                "B170": parsed_data["sec9"].get("B170","").capitalize(),
                                "B176": parsed_data["sec9"].get("B176",""),
                                "B183": parsed_data["sec9"].get("B183",""),
                                "B189": parsed_data["sec9"].get("B189","")
                            }
                            
                            for addr, val in cell_map_e.items():
                                if not val: continue
                                if addr in ["B183", "B189"] and "±" not in val:
                                    num = re.search(r'([\d\.]+)', val)
                                    if num:
                                        suffix = "0.01" if addr == "B183" else "0.005"
                                        val = f"{num.group(1)} ± {suffix}"
                                
                                formatted, h = format_and_calc_height_sec47(val)
                                r_idx = int(re.search(r'\d+', addr).group())
                                safe_write_force(dest_ws, r_idx, 2, formatted, center=False)
                                dest_ws.row_dimensions[r_idx].height = h

                            s8 = parsed_data["sec8"]
                            if s8["B154"]:
                                lines = s8["B154"].split('\n')
                                safe_write_force(dest_ws, 154, 2, lines[0].lower() if "no data" in lines[0].lower() else lines[0], center=False)
                                if len(lines) > 1:
                                    safe_write_force(dest_ws, 155, 2, "\n".join(lines[1:]), center=False)
                                    dest_ws.row_dimensions[155].hidden = False
                            
                            if s8["B156"]:
                                lines = s8["B156"].split('\n')
                                safe_write_force(dest_ws, 156, 2, lines[0].lower() if "no data" in lines[0].lower() else lines[0], center=False)
                                if len(lines) > 1:
                                    safe_write_force(dest_ws, 157, 2, "\n".join(lines[1:]), center=False)
                                    dest_ws.row_dimensions[157].hidden = False

                            fill_regulatory_section(dest_ws, 202, 240, active_substances, eng_data_map, 'F')
                            fill_regulatory_section(dest_ws, 242, 279, active_substances, eng_data_map, 'G')
                            fill_regulatory_section(dest_ws, 281, 315, active_substances, eng_data_map, 'H')
                            fill_regulatory_section(dest_ws, 324, 358, active_substances, eng_data_map, 'P')
                            fill_regulatory_section(dest_ws, 360, 395, active_substances, eng_data_map, 'Q')
                            fill_regulatory_section(dest_ws, 401, 437, active_substances, eng_data_map, 'T')
                            fill_regulatory_section(dest_ws, 439, 478, active_substances, eng_data_map, 'U')
                            fill_regulatory_section(dest_ws, 480, 519, active_substances, eng_data_map, 'V')

                            s14 = parsed_data["sec14"]
                            safe_write_force(dest_ws, 531, 2, s14["UN"], center=False)
                            name_clean = re.sub(r'[\(\)\d]', '', s14["NAME"]).strip()
                            safe_write_force(dest_ws, 532, 2, name_clean, center=False)

                            today_eng = datetime.now().strftime("%d. %b. %Y")
                            safe_write_force(dest_ws, 544, 1, f"16.2 Date of Issue : {today_eng}", center=False)

                        else: # CFF(K) / HP(K)
                            safe_write_force(dest_ws, 7, 2, product_name_input, center=True)
                            safe_write_force(dest_ws, 10, 2, product_name_input, center=True)
                            
                            if parsed_data["hazard_cls"]:
                                clean_hazard_text = "\n".join([line for line in parsed_data["hazard_cls"] if line.strip()])
                                safe_write_force(dest_ws, 20, 2, clean_hazard_text, center=False)
                                dest_ws['B20'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')

                            signal_final = parsed_data["signal_word"] if parsed_data["signal_word"] else ""
                            safe_write_force(dest_ws, 24, 2, signal_final, center=False) 

                            if option == "HP(K)":
                                safe_write_force(dest_ws, 38, 1, "예방", center=False)
                                safe_write_force(dest_ws, 50, 1, "대응", center=False)
                                safe_write_force(dest_ws, 64, 1, "저장", center=False)
                                safe_write_force(dest_ws, 70, 1, "폐기", center=False)

                            fill_fixed_range(dest_ws, 25, 36, parsed_data["h_codes"], code_map)
                            fill_fixed_range(dest_ws, 38, 49, parsed_data["p_prev"], code_map)
                            fill_fixed_range(dest_ws, 50, 63, parsed_data["p_resp"], code_map)
                            fill_fixed_range(dest_ws, 64, 69, parsed_data["p_stor"], code_map)
                            fill_fixed_range(dest_ws, 70, 72, parsed_data["p_disp"], code_map)

                            fill_composition_data(dest_ws, parsed_data["composition_data"], cas_name_map, mode=option)
                            
                            active_substances = []
                            for c_data in parsed_data["composition_data"]:
                                cas = c_data[0].replace(" ", "").strip()
                                if cas in cas_name_map:
                                    name = cas_name_map[cas]
                                    if name: active_substances.append(name)

                            sec_data = parsed_data["sec4_to_7"]
                            for cell_addr, raw_text in sec_data.items():
                                formatted_txt, row_h = format_and_calc_height_sec47(raw_text)
                                try:
                                    col_str = re.match(r"([A-Z]+)", cell_addr).group(1)
                                    row_num = int(re.search(r"(\d+)", cell_addr).group(1))
                                    col_idx = openpyxl.utils.column_index_from_string(col_str)
                                    safe_write_force(dest_ws, row_num, col_idx, "")
                                    if formatted_txt:
                                        safe_write_force(dest_ws, row_num, col_idx, formatted_txt, center=False)
                                        dest_ws.row_dimensions[row_num].height = row_h
                                        try:
                                            cell_a = dest_ws.cell(row=row_num, column=1)
                                            if cell_a.value: cell_a.value = str(cell_a.value).strip()
                                            cell_a.alignment = ALIGN_TITLE
                                        except: pass
                                except Exception as e: pass

                            s8 = parsed_data["sec8"]
                            val148 = s8["B148"].replace("해당없음", "자료없음")
                            lines148 = [l.strip() for l in val148.split('\n') if l.strip()]
                            safe_write_force(dest_ws, 148, 2, ""); safe_write_force(dest_ws, 149, 2, ""); dest_ws.row_dimensions[149].hidden = True
                            if lines148:
                                safe_write_force(dest_ws, 148, 2, lines148[0], center=False)
                                if len(lines148) > 1:
                                    safe_write_force(dest_ws, 149, 2, "\n".join(lines148[1:]), center=False)
                                    dest_ws.row_dimensions[149].hidden = False
                            
                            val150 = s8["B150"].replace("해당없음", "자료없음")
                            val150 = re.sub(r"^규정[:\s]*", "", val150).strip()
                            safe_write_force(dest_ws, 150, 2, val150, center=False)

                            s9 = parsed_data["sec9"]
                            safe_write_force(dest_ws, 163, 2, s9["B163"], center=False)
                            
                            if option == "HP(K)":
                                flash = s9["B169"]
                                flash_num = re.findall(r'([<>]?\s*\d{2,3})', flash)
                                safe_write_force(dest_ws, 169, 2, f"{flash_num[0]}℃" if flash_num else "", center=False)
                            else:
                                flash = s9["B169"]
                                flash_num = re.findall(r'(\d{2,3})', flash)
                                safe_write_force(dest_ws, 169, 2, f"{flash_num[0]}℃" if flash_num else "", center=False)
                            
                            gravity = s9["B176"].replace("(20℃)", "").replace("(물=1)", "")
                            g_match = re.search(r'([\d\.]+)', gravity)
                            safe_write_force(dest_ws, 176, 2, f"{g_match.group(1)} ± 0.01" if g_match else "", center=False)
                            
                            refract = s9["B182"].replace("(20℃)", "")
                            r_match = re.search(r'([\d\.]+)', refract)
                            safe_write_force(dest_ws, 182, 2, f"{r_match.group(1)} ± 0.005" if r_match else "", center=False)

                            fill_regulatory_section(dest_ws, 195, 226, active_substances, kor_data_map, 'F')
                            fill_regulatory_section(dest_ws, 228, 260, active_substances, kor_data_map, 'G')
                            fill_regulatory_section(dest_ws, 269, 300, active_substances, kor_data_map, 'H')
                            fill_regulatory_section(dest_ws, 316, 348, active_substances, kor_data_map, 'P')
                            fill_regulatory_section(dest_ws, 353, 385, active_substances, kor_data_map, 'P')
                            fill_regulatory_section(dest_ws, 392, 426, active_substances, kor_data_map, 'T')
                            fill_regulatory_section(dest_ws, 428, 460, active_substances, kor_data_map, 'U')
                            fill_regulatory_section(dest_ws, 465, 497, active_substances, kor_data_map, 'V')

                            for r in range(261, 268): dest_ws.row_dimensions[r].hidden = True
                            for r in range(349, 352): dest_ws.row_dimensions[r].hidden = True
                            dest_ws.row_dimensions[386].hidden = True
                            for r in range(461, 464): dest_ws.row_dimensions[r].hidden = True

                            s14 = parsed_data["sec14"]
                            un_val = re.sub(r"\D", "", s14["UN"])
                            safe_write_force(dest_ws, 512, 2, un_val, center=False)
                            
                            name_val = re.sub(r"\([^)]*\)", "", s14["NAME"]).strip()
                            safe_write_force(dest_ws, 513, 2, name_val, center=False)

                            s15 = parsed_data["sec15"]
                            if option == "CFF(K)":
                                safe_write_force(dest_ws, 521, 2, s15["DANGER"], center=False)

                            today_str = datetime.now().strftime("%Y.%m.%d")
                            safe_write_force(dest_ws, 542, 2, today_str, center=False)

                        # [공통] 이미지 처리 - HP(K) 복원
                        collected_pil_images = []
                        page = doc[0]
                        image_list = doc.get_page_images(0)
                        
                        for img_info in image_list:
                            xref = img_info[0]
                            # HP(K) 필터: 상단 로고, 파란색, 정사각형 아님 제거
                            if option == "HP(K)":
                                try:
                                    rect = page.get_image_bbox(img_info)
                                    if rect.y1 < (page.rect.height * 0.15): continue
                                    width = rect.x1 - rect.x0; height = rect.y1 - rect.y0
                                    if not is_square_shaped(width, height): continue
                                except: continue

                            try:
                                base_image = doc.extract_image(xref)
                                pil_img = PILImage.open(io.BytesIO(base_image["image"]))
                                
                                if option == "HP(K)":
                                    if is_blue_dominant(pil_img): continue

                                if loaded_refs:
                                    # [HP(K) 복원] 점수(score)도 함께 받아옴
                                    matched_name, score = find_best_match_name(pil_img, loaded_refs, mode=option)
                                    if matched_name:
                                        clean_img = loaded_refs[matched_name]
                                        collected_pil_images.append((extract_number(matched_name), clean_img, score))
                            except: continue
                        
                        # [HP(K) 복원] 점수 기반 필터링 및 중복 제거
                        final_images_map = {}
                        if option == "HP(K)" and collected_pil_images:
                            min_score = min(item[2] for item in collected_pil_images)
                            for key, img, score in collected_pil_images:
                                if score > min_score + 25: continue # 1등과 차이 많이 나면 버림
                                if key not in final_images_map: final_images_map[key] = (img, score)
                                else:
                                    if score < final_images_map[key][1]: final_images_map[key] = (img, score)
                        else:
                            for key, img, _ in collected_pil_images:
                                if key not in final_images_map: final_images_map[key] = (img, 0)
                        
                        # [HP(K) 복원] 이미지 객체만 추출 (item[1]이 (img, score) 튜플일 수 있으므로 처리)
                        final_sorted_imgs = []
                        for item in sorted(final_images_map.items(), key=lambda x: x[0]):
                            val = item[1]
                            if isinstance(val, tuple): # (img, score) 형태면 img만
                                final_sorted_imgs.append(val[0])
                            else: # img만 있으면 그대로
                                final_sorted_imgs.append(val)

                        if final_sorted_imgs:
                            unit_size = 67; icon_size = 60
                            padding_top = 4; padding_left = (unit_size - icon_size) // 2
                            total_width = unit_size * len(final_sorted_imgs)
                            total_height = unit_size
                            merged_img = PILImage.new('RGBA', (total_width, total_height), (255, 255, 255, 0))
                            for idx, p_img in enumerate(final_sorted_imgs):
                                p_img_resized = p_img.resize((icon_size, icon_size), PILImage.LANCZOS)
                                merged_img.paste(p_img_resized, ((idx * unit_size) + padding_left, padding_top))
                            
                            img_byte_arr = io.BytesIO()
                            merged_img.save(img_byte_arr, format='PNG')
                            img_byte_arr.seek(0)
                            # CFF(E)는 B22, 나머지는 B23
                            dest_ws.add_image(XLImage(img_byte_arr), 'B22' if option=="CFF(E)" else 'B23') 

                        dest_wb.external_links = []
                        output = io.BytesIO()
                        dest_wb.save(output)
                        output.seek(0)
                        
                        final_name = f"{product_name_input} GHS MSDS({'E' if 'E' in option else 'K'}).xlsx"
                        if final_name in new_download_data:
                            final_name = f"{product_name_input}_{uploaded_file.name.split('.')[0]}.xlsx"
                        new_download_data[final_name] = output.getvalue()
                        new_files.append(final_name)
                        
                    except Exception as e:
                        st.error(f"오류 ({uploaded_file.name}): {e}")

            st.session_state['converted_files'] = new_files
            st.session_state['download_data'] = new_download_data
            
            if 'df_code' in locals(): del df_code
            if 'doc' in locals(): doc.close()
            if 'dest_wb' in locals(): del dest_wb
            if 'output' in locals(): del output
            gc.collect()

            if new_files: st.success("완료!")
    else:
        st.error("모든 파일을 업로드해주세요.")

with col_right:
    st.subheader("결과 다운로드")
    if st.session_state['converted_files']:
        for i, fname in enumerate(st.session_state['converted_files']):
            c1, c2 = st.columns([3, 1])
            with c1: st.text(f"📄 {fname}")
            with c2:
                st.download_button(
                    label="받기", 
                    data=st.session_state['download_data'][fname], 
                    file_name=fname, 
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=i
                )
